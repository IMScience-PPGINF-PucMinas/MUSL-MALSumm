# -*- coding: utf-8 -*-
import torch
from os import listdir
import numpy as np
from os.path import join
import h5py
import json
import argparse
import re
import os
import logging
from utils.utils import get_paths, setup_logging
from evaluation.evaluation_metrics import evaluate_summary
from model.layers.summarizer import xLSTM
from inference.generate_summary import generate_summary, expand_scores_to_frames
from scipy.stats import kendalltau, spearmanr

setup_logging()


# ---------------------------------------------------------------------------
# Device selection
# ---------------------------------------------------------------------------

def _get_device():
    """Return the best available torch device (CUDA > CPU)."""
    return torch.device("cuda" if torch.cuda.is_available() else "cpu")


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def load_video_data(dataset, hdf, video):
    """Load video data for a single video from an already-open HDF5 handle.

    Accepting the open handle (rather than the file path) allows the caller to
    keep the file open across all videos in a split, avoiding repeated open/close
    overhead that is measurable on datasets with many videos.

    :param str dataset: Dataset name (SumMe | TVSum | MrHiSum).
    :param h5py.File hdf: Open HDF5 file handle (read mode).
    :param str video: HDF5 group key for the target video.
    :return: tuple (frame_features, user_summary, sb, n_frames, positions, video_name)
    """
    frame_features = torch.Tensor(
        np.array(hdf[f"{video}/features"])
    ).view(-1, 1024)
    sb         = np.array(hdf[f"{video}/change_points"])
    video_name = None

    if dataset.lower() in ('summe', 'tvsum'):
        user_summary = np.array(hdf[f"{video}/user_summary"])
        n_frames     = np.array(hdf[f"{video}/n_frames"])
        positions    = np.array(hdf[f"{video}/picks"])
        if "video_name" in hdf[f"{video}"]:
            video_name = str(
                np.array(hdf[f"{video}/video_name"]).astype(str, copy=False)
            )
    elif dataset.lower() == 'mrhisum':
        user_summary = np.array(hdf[f"{video}/gt_summary"])
        n_frames     = frame_features.shape[0]
        positions    = np.arange(n_frames, dtype=int)
    else:
        raise ValueError(f"Unsupported dataset: {dataset}")

    return frame_features, user_summary, sb, n_frames, positions, video_name


# ---------------------------------------------------------------------------
# Best-epoch selection helpers
# ---------------------------------------------------------------------------

def _find_epoch_files(model_path):
    """Return all epoch-N.pkl files in model_path, sorted by epoch number."""
    files = [
        f for f in listdir(model_path)
        if re.match(r"epoch-\d+\.pkl", f)
    ]
    return sorted(files, key=lambda x: int(re.findall(r'\d+', x)[0]))


def _load_best_epoch_from_fscores(model_path):
    """Read pre-computed f_scores.txt produced by compute_fscores.py.

    Returns the best epoch number (0-indexed line = 0-indexed epoch).
    Returns None if the file does not exist.
    """
    fscores_path = join(model_path, 'f_scores.txt')
    if not os.path.exists(fscores_path):
        return None
    with open(fscores_path) as fp:
        content = fp.read().strip()
    try:
        scores = json.loads(content)
    except json.JSONDecodeError:
        scores = [float(x) for x in content.splitlines()]
    return int(np.argmax(scores))


# ---------------------------------------------------------------------------
# Core inference
# ---------------------------------------------------------------------------

def run_inference(model, data_path, keys, eval_method, save_summary,
                  dataset, device=None, verbose=False):
    """Run inference for a single model checkpoint over all test videos.

    The HDF5 file is opened once for the entire split and reused across
    videos, eliminating per-video open/close overhead.

    The model is moved to *device* before inference; features are moved
    to the same device so that the forward pass executes on GPU when available.

    frame_scores for Kendall/Spearman are now computed via expand_scores_to_frames
    (shared with generate_summary) instead of duplicating the expansion logic.

    Returns:
        mean_fscore, mean_kendall, mean_spearman,
        video_summaries, [video_names if SumMe]
    """
    if device is None:
        device = next(model.parameters()).device

    model.eval()

    video_fscores   = []
    video_kendalls  = []
    video_spearmans = []
    video_summaries = {}
    video_names     = {}
    summe = (dataset.lower() == 'summe')

    # Single HDF5 open for the entire split
    with h5py.File(data_path, "r") as hdf:
        for video in keys:
            if summe:
                try:
                    if int(video.split('_')[1]) > 25:
                        continue
                except (IndexError, ValueError):
                    pass

            frame_features, user_summary, sb, n_frames, positions, vname = \
                load_video_data(dataset, hdf, video)

            # Move features to GPU (no-op if device is CPU)
            frame_features = frame_features.to(device)

            with torch.no_grad():
                scores, _ = model(frame_features)
                scores = scores.squeeze(0).cpu().numpy().tolist()

            summary = generate_summary([sb], [scores], [n_frames], [positions])[0]
            f_score = evaluate_summary(summary, user_summary, eval_method)

            # Expand sub-sampled scores to full frame sequence (shared utility —
            # no duplication of the propagation logic that lives in generate_summary)
            frame_scores = expand_scores_to_frames(
                np.array(scores), positions, int(n_frames)
            )

            gt_importance = (
                user_summary.mean(axis=0) if user_summary.ndim > 1 else user_summary
            )

            if frame_scores.shape[0] != gt_importance.shape[0]:
                logging.warning(
                    f"Shape mismatch for {video}: "
                    f"pred={frame_scores.shape[0]}, gt={gt_importance.shape[0]}"
                    " — skipping correlations"
                )
                ktau, spr = float('nan'), float('nan')
            else:
                ktau, _ = kendalltau(frame_scores, gt_importance)
                spr,  _ = spearmanr(frame_scores,  gt_importance)

            video_fscores.append(f_score)
            video_kendalls.append(ktau)
            video_spearmans.append(spr)
            video_summaries[video] = summary

            if summe:
                video_names[video] = vname

            if verbose:
                logging.info(
                    f"  {video} ({vname}): F1={f_score:.2f}%  τ={ktau:.4f}  ρ={spr:.4f}"
                )

            if save_summary:
                out   = {str(i): int(v) for i, v in enumerate(summary)}
                fname = f"{video}_summary.json"
                with open(fname, "w") as fp:
                    json.dump(out, fp, indent=4)
                print(f"Summary saved → {fname}")

    mean_fscore   = float(np.nanmean(video_fscores))
    mean_kendall  = float(np.nanmean(video_kendalls))
    mean_spearman = float(np.nanmean(video_spearmans))

    if summe:
        return mean_fscore, mean_kendall, mean_spearman, video_summaries, video_names
    return mean_fscore, mean_kendall, mean_spearman, video_summaries


# ---------------------------------------------------------------------------
# Parallel full-scan worker
# ---------------------------------------------------------------------------

def _scan_split_worker(args):
    """Evaluate all epoch checkpoints for a single split.

    Designed to run inside a separate process via ProcessPoolExecutor.
    All arguments are plain Python objects (pickle-safe) — no live model or
    tensor objects are passed across the process boundary.

    The model is instantiated once per split and reused across all epochs:
    only load_state_dict is called per epoch, avoiding repeated constructor and
    memory allocation overhead.

    GPU is used when available. In multi-process mode each worker claims the
    same default CUDA device; if multiple GPUs are present, assign
    CUDA_VISIBLE_DEVICES externally or pass a device index via model_kwargs.

    Args:
        args: tuple of
            (split_id, model_path, epoch_files, dataset_path,
             test_keys, eval_metric, dataset, model_kwargs, verbose)

    Returns:
        (split_id, best_epoch, results_dict)
        where results_dict maps epoch_num → (fscore, kendall, spearman).
    """
    (split_id, model_path, epoch_files,
     dataset_path, test_keys,
     eval_metric, dataset, model_kwargs, verbose) = args

    device = _get_device()

    # Instantiate the model once and reuse across all epochs
    model = xLSTM(**model_kwargs).to(device)
    model.eval()

    results = {}
    for fname in epoch_files:
        epoch_num = int(re.findall(r'\d+', fname)[0])

        # Swap weights only — no reallocation
        state_dict = torch.load(join(model_path, fname), map_location=device)
        model.load_state_dict(state_dict)

        fs, kt, sp, *_ = run_inference(
            model, dataset_path, test_keys,
            eval_metric, save_summary=False,
            dataset=dataset, device=device, verbose=verbose,
        )
        results[epoch_num] = (fs, kt, sp)

    best_epoch = max(results, key=lambda e: results[e][0])
    return split_id, best_epoch, results


def _run_full_scan_parallel(split_ids, split_configs, n_workers):
    """Run full epoch scan for all splits in parallel.

    Uses 'spawn' as the multiprocessing start method so that CUDA can be
    initialized cleanly inside each worker. The default 'fork' start method
    on Linux causes "Cannot re-initialize CUDA in forked subprocess" because
    the CUDA context created in the parent process is inherited but unusable
    by the child. 'spawn' starts a fresh interpreter in each worker, avoiding
    this entirely.

    Args:
        split_ids:    list of split indices to process
        split_configs: dict mapping split_id → worker args tuple
        n_workers:    number of parallel worker processes

    Returns:
        dict mapping split_id → (best_epoch, results_dict)
    """
    import multiprocessing as mp

    n_workers = min(n_workers, len(split_ids))
    print(f"Full scan: {n_workers} parallel worker(s) across {len(split_ids)} splits\n")

    output = {}
    ctx    = mp.get_context("spawn")

    with ctx.Pool(processes=n_workers) as pool:
        async_results = {
            s: pool.apply_async(_scan_split_worker, (split_configs[s],))
            for s in split_ids
        }

        for split_id, ar in async_results.items():
            try:
                sid, best_epoch, results = ar.get()
                output[sid] = (best_epoch, results)
                best_fs = results[best_epoch][0]
                print(
                    f"  Split {sid} done — best epoch: {best_epoch} "
                    f"(F1={best_fs:.2f}%)"
                )
            except Exception as exc:
                logging.error(f"  Split {split_id} failed: {exc}")

    return output


# ---------------------------------------------------------------------------
# Output helpers
# ---------------------------------------------------------------------------

def _print_results(split_ids, best_epochs, split_results,
                   best_avg_epoch, avg_fs, avg_ks, avg_ss):
    """Print a clean summary table to stdout."""
    sep = "-" * 62
    print(f"\n{sep}")
    print(
        f"{'Split':<8} {'Epoch':>6}  {'F1 (%)':>8}  "
        f"{'Kendall τ':>10}  {'Spearman ρ':>11}"
    )
    print(sep)
    for s in split_ids:
        if s not in best_epochs:
            continue
        ep       = best_epochs[s]
        fs, kt, sp = split_results[s]
        print(f"  {s:<6} {ep:>6}  {fs:>8.2f}  {kt:>10.4f}  {sp:>11.4f}")
    print(sep)
    print(
        f"  {'AVG':<6} {best_avg_epoch:>6}  "
        f"{avg_fs[best_avg_epoch]:>8.2f}  "
        f"{avg_ks[best_avg_epoch]:>10.4f}  "
        f"{avg_ss[best_avg_epoch]:>11.4f}"
    )
    print(f"{sep}\n")


def _save_xlsx(split_ids, all_epoch_results, dataset):
    """Save full per-epoch metrics to an xlsx file.

    Only called when --save_results=1. pandas/openpyxl are imported here
    so they add zero overhead in the default fast path.
    """
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment

    all_epochs = sorted({
        ep
        for s in split_ids
        for ep in all_epoch_results.get(s, {})
    })

    rows       = {"Epoch": all_epochs}
    avg_fs, avg_ks, avg_ss = {}, {}, {}

    for ep in all_epochs:
        vf = [all_epoch_results[s][ep][0] for s in split_ids if ep in all_epoch_results.get(s, {})]
        vk = [all_epoch_results[s][ep][1] for s in split_ids if ep in all_epoch_results.get(s, {})]
        vs = [all_epoch_results[s][ep][2] for s in split_ids if ep in all_epoch_results.get(s, {})]
        avg_fs[ep] = float(np.nanmean(vf)) if vf else float('nan')
        avg_ks[ep] = float(np.nanmean(vk)) if vk else float('nan')
        avg_ss[ep] = float(np.nanmean(vs)) if vs else float('nan')

    for s in split_ids:
        er = all_epoch_results.get(s, {})
        rows[f"F-score Split {s}"]  = [er.get(ep, (None,))[0]          for ep in all_epochs]
        rows[f"Kendall Split {s}"]  = [er.get(ep, (None, None))[1]      for ep in all_epochs]
        rows[f"Spearman Split {s}"] = [er.get(ep, (None, None, None))[2] for ep in all_epochs]
    rows["Avg F-score"]  = [avg_fs[ep] for ep in all_epochs]
    rows["Avg Kendall"]  = [avg_ks[ep] for ep in all_epochs]
    rows["Avg Spearman"] = [avg_ss[ep] for ep in all_epochs]

    df = pd.DataFrame(rows).set_index("Epoch")

    tuples = []
    for s in split_ids:
        for m in ("F-score", "Kendall", "Spearman"):
            tuples.append((f"Split {s}", m))
    for m in ("F-score", "Kendall", "Spearman"):
        tuples.append(("Average", m))
    df.columns = pd.MultiIndex.from_tuples(tuples)

    xlsx_path = f"{dataset}_epoch_metrics.xlsx"
    df.to_excel(xlsx_path)

    wb = load_workbook(xlsx_path)
    ws = wb.active
    ws.merge_cells("A1:A2")
    cell       = ws["A1"]
    cell.value = "Epoch"
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.delete_rows(3, 1)
    wb.save(xlsx_path)

    print(f"Full epoch metrics saved → {xlsx_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Run inference and report best-epoch results for each split."
    )
    parser.add_argument("--dataset",       type=str,   default='SumMe',
                        help="Dataset [SumMe | TVSum | MrHiSum]")
    parser.add_argument("--model_version", type=str,   default='',
                        help="Model version suffix, e.g. 'v2'")
    parser.add_argument("--verbose",       type=int,   default=0,
                        help="Per-video log (0=off, 1=on)")
    parser.add_argument("--save_summary",  type=int,   default=0,
                        help="Export binary summary JSON per video (0=off, 1=on)")
    parser.add_argument("--save_results",  type=int,   default=0,
                        help="Full epoch scan + save xlsx (0=off, 1=on)")
    parser.add_argument("--workers",       type=int,   default=5,
                        help="Parallel worker processes for full scan "
                             "(--save_results=1 only). Default=5 (one per split). "
                             "Set to 1 to disable parallelism.")
    parser.add_argument("--hidden_dim",    type=int,   default=512)
    parser.add_argument("--num_layers",    type=int,   default=2)
    parser.add_argument("--dropout",       type=float, default=0.5,
                        help="Must match the value used during training.")

    args = vars(parser.parse_args())

    dataset       = args["dataset"]
    model_version = args["model_version"]
    verbose       = bool(args["verbose"])
    save_summary  = bool(args["save_summary"])
    save_results  = bool(args["save_results"])
    n_workers     = args["workers"]

    eval_metric = 'avg' if dataset.lower() == 'tvsum' else 'max'
    split_ids   = list(range(5)) if dataset.lower() in ('summe', 'tvsum') else [0]

    model_kwargs = dict(
        input_size=1024,
        output_size=1024,
        num_segments=4,
        hidden_dim=args["hidden_dim"],
        num_layers=args["num_layers"],
        dropout=args["dropout"],
    )

    paths        = get_paths(dataset)
    dataset_path = paths['dataset']
    split_file   = paths['split']

    with open(split_file) as fp:
        split_data = json.load(fp)

    device = _get_device()
    print(f"\nDataset: {dataset}  |  eval: {eval_metric}  |  splits: {split_ids}"
          f"  |  device: {device}")

    # -----------------------------------------------------------------------
    # Full scan path (--save_results 1): parallel processing across splits
    # -----------------------------------------------------------------------
    if save_results:
        split_configs = {}
        for split_id in split_ids:
            model_path = (
                f"Summaries/xLSTM/{dataset}{model_version}/models/split{split_id}"
            )
            test_keys = (
                split_data[split_id]["test_keys"]
                if isinstance(split_data, list)
                else split_data["test_keys"]
            )
            epoch_files = _find_epoch_files(model_path)

            if not epoch_files:
                logging.warning(
                    f"No epoch files in {model_path} — skipping split {split_id}"
                )
                continue

            split_configs[split_id] = (
                split_id, model_path, epoch_files,
                dataset_path, test_keys,
                eval_metric, dataset, model_kwargs, verbose,
            )

        if not split_configs:
            print("No valid splits found. Check model paths.")
            return

        scan_output = _run_full_scan_parallel(
            list(split_configs.keys()), split_configs, n_workers
        )

        best_epochs       = {}
        split_results     = {}
        all_epoch_results = {}

        for sid, (best_epoch, results) in scan_output.items():
            best_epochs[sid]       = best_epoch
            split_results[sid]     = results[best_epoch]
            all_epoch_results[sid] = results

    # -----------------------------------------------------------------------
    # Fast path (--save_results 0): one inference per split, GPU-accelerated
    # -----------------------------------------------------------------------
    else:
        best_epochs   = {}
        split_results = {}

        for split_id in split_ids:
            model_path = (
                f"Summaries/xLSTM/{dataset}{model_version}/models/split{split_id}"
            )
            test_keys = (
                split_data[split_id]["test_keys"]
                if isinstance(split_data, list)
                else split_data["test_keys"]
            )
            epoch_files = _find_epoch_files(model_path)

            if not epoch_files:
                logging.warning(
                    f"No epoch files in {model_path} — skipping split {split_id}"
                )
                continue

            best_epoch = _load_best_epoch_from_fscores(model_path)

            if best_epoch is not None:
                best_pkl = join(model_path, 'best_model.pkl')
                fname    = (
                    'best_model.pkl'
                    if os.path.exists(best_pkl)
                    else f'epoch-{best_epoch}.pkl'
                )
                print(f"Split {split_id}: epoch {best_epoch} (from f_scores.txt)")
            else:
                fname      = epoch_files[-1]
                best_epoch = int(re.findall(r'\d+', fname)[0])
                print(
                    f"Split {split_id}: f_scores.txt not found — "
                    f"using last epoch ({best_epoch})"
                )

            model = xLSTM(**model_kwargs)
            model.load_state_dict(
                torch.load(join(model_path, fname), map_location=device)
            )
            model = model.to(device)

            fs, kt, sp, *_ = run_inference(
                model, dataset_path, test_keys,
                eval_metric, save_summary, dataset,
                device=device, verbose=verbose,
            )
            best_epochs[split_id]   = best_epoch
            split_results[split_id] = (fs, kt, sp)

    # -----------------------------------------------------------------------
    # Print consolidated results
    # -----------------------------------------------------------------------
    if not split_results:
        print("No results collected — check model paths and split files.")
        return

    valid_splits   = list(split_results.keys())
    all_f = [split_results[s][0] for s in valid_splits]
    all_k = [split_results[s][1] for s in valid_splits]
    all_s = [split_results[s][2] for s in valid_splits]

    best_avg_epoch = best_epochs[max(valid_splits, key=lambda s: split_results[s][0])]
    avg_fs = {best_avg_epoch: float(np.nanmean(all_f))}
    avg_ks = {best_avg_epoch: float(np.nanmean(all_k))}
    avg_ss = {best_avg_epoch: float(np.nanmean(all_s))}

    _print_results(
        valid_splits, best_epochs, split_results,
        best_avg_epoch, avg_fs, avg_ks, avg_ss,
    )

    if save_results:
        _save_xlsx(valid_splits, all_epoch_results, dataset)


if __name__ == "__main__":
    main()